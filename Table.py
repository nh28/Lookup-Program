import logging
import pandas as pd
import openpyxl

class Table:
    def __init__(self, arkeon, user_station, user_elements, user_months):
        """
        Initialize a new instance of Table.

        Parameters:
        arkeon: An instance of the ARKEON object to be able to query data.
        user_station: The station name that the user selected.
        user_elements: A list of elements that the user selected.
        user_months: A list of months that the user selected.
        """
        self.headers = ['STN ID', 'STN/LOC NAME', 'CLIMATE ID', 'PROV', 'NORMAL ID', 'ELEMENT NAME', 'Month', 'Value (71)', 'Code (71)', 'Date (71)', 'Value (81)', 'Code (81)', 'Date (81)', 'Value (91)', 'Code (91)', 'Date (91)']
        self.arkeon = arkeon
        self.user_station = user_station
        self.user_elements = user_elements
        self.user_months = user_months
        self.extreme_el = self.arkeon.get_extreme_el()
        try:
            self.workbook = openpyxl.load_workbook('StationList.xlsx')
            self.worksheet = self.workbook["Sheet1"]
        except Exception as e:
            logging.error("Unable to find StationList.xlsx with Sheet1")
            print(e)

    def check_station_availability(self):
        """
        Finds the index in StationList.xlsx of the station that the user selected.

        Parameters:
        None

        Returns:
        row_idx: The index of the station list, and returns None if the station is not found.
        """
        for row_idx in range(2, self.worksheet.max_row):
            name_7181 = self.worksheet.cell(row = row_idx, column = 2).value
            name_91 = self.worksheet.cell(row = row_idx, column = 7).value
            if name_7181 == self.user_station or name_91 == self.user_station:
                return row_idx
        return None

    def format(self, extreme, df):
        """
        Formats the data based on the dataframe and if the element is an extreme element.

        Parameters:
        extreme: True if the given element is an extreme element.
        df: A dataframe that contains the data based on station, element, and month

        Returns:
        A formatted list with the information.
        """
        if df.empty:
            return [ "", "", ""]
        elif extreme:
            return [ df.iloc[0]["VALUE"],  df.iloc[0]["NORMAL_CODE"],  df.iloc[0]["FIRST_OCCURRENCE_DATE"]]
        else:
            return [ df.iloc[0]["VALUE"],  df.iloc[0]["NORMAL_CODE"],  ""]

    def get_all_data(self):
        """
        Loops through all the user elements and months to compile a dataframe containing all the data.

        Parameters:
        None

        Returns: 
        df: A dataframe that contains all the information that can be inputted into a table.
        """
        df = pd.DataFrame(columns=self.headers)
        df_81 = pd.DataFrame(columns=self.headers)
        df_91 = pd.DataFrame(columns=self.headers)
        
        row_idx = self.check_station_availability()
        if row_idx != None:
            station_info = [cell.value for cell in self.worksheet[row_idx]]
            id_7181 = station_info[0]
            name_7181 = station_info[1]
            climate_id = station_info[2]
            exist_71 = station_info[3]
            exist_81 = station_info[4]
            id_91 = station_info[5]
            name_91 = station_info[6]
            prov = station_info[7]

            for element in self.user_elements:
                for month in self.user_months:
                    normal_id = self.arkeon.get_normals_element_id(element)
                    if element in self.extreme_el:
                        extreme = True
                    else:
                        extreme = False
                    
                    if exist_71 != None and exist_71 != "":
                        query_71 = self.arkeon.get_dataframe(['VALUE', 'FIRST_OCCURRENCE_DATE', 'NORMAL_CODE'], 'NORMALS.NORMALS_DATA', ['stn_id = ' + str(id_7181), 'normal_id = ' + str(normal_id), 'month = ' + str(month)], True)
                        data_71 = self.format(extreme, query_71)
                    if exist_81 != None and exist_81 != "":
                        query_81 = self.arkeon.get_dataframe(['VALUE', 'FIRST_OCCURRENCE_DATE', 'NORMAL_CODE'], 'NORMALS_1981.NORMALS_DATA', ['stn_id = ' + str(id_7181), 'normal_id = ' + str(normal_id), 'month = ' + str(month)], True)
                        data_81 = self.format(extreme, query_81)
                    if id_91 != None and id_91 != "":
                        query_91 = self.arkeon.get_dataframe(['VALUE', 'FIRST_OCCURRENCE_DATE', 'NORMAL_CODE'], 'NORMALS_1991.NORMALS_DATA', ['stn_id = ' + str(id_91), 'normal_id = ' + str(normal_id), 'month = ' + str(month)], True)
                        data_91 = self.format(extreme, query_91)


                    if exist_71 != None and exist_81 != None and exist_71 != "" and exist_81 != "": 
                        if id_91 != None and id_91 != "":
                            # might need to add a statement here about if the stations match or not
                            row = [id_7181, name_7181, climate_id, prov, normal_id, element, month] + data_71 + data_81 + data_91
                            row_df = pd.DataFrame([row], columns = self.headers)
                            df = pd.concat([df, row_df], ignore_index=False)
                        else:
                            row = [id_7181, name_7181, climate_id, prov, normal_id, element, month] + data_71 + data_81 + ["","",""]
                            row_df = pd.DataFrame([row], columns = self.headers)
                            df = pd.concat([df, row_df], ignore_index=False)
                    elif exist_71 != None and exist_71 != "":
                        if id_91 != None and id_91 != "":
                            row = [id_7181, name_7181, "", prov, normal_id, element, month] + data_71 + ["","",""] + data_91
                            row_df = pd.DataFrame([row], columns = self.headers)
                            df = pd.concat([df, row_df], ignore_index=False)
                        else:
                            row = [id_7181, name_7181, "", prov, normal_id, element, month] + data_71 + ["","",""] + ["","",""]
                            row_df = pd.DataFrame([row], columns = self.headers)
                            df = pd.concat([df, row_df], ignore_index=False)
                    elif exist_81 != None and exist_81 != "":
                        if id_91 != None and id_91 != "":
                            row = [id_7181, name_7181, climate_id, prov, normal_id, element, month] + ["","",""] + data_81 + data_91
                            row_df = pd.DataFrame([row], columns = self.headers)
                            df_81 = pd.concat([df_81, row_df], ignore_index=False)
                        else:
                            row = [id_7181, name_7181, climate_id, prov, normal_id, element, month] + ["","",""] + data_81 + ["","",""]
                            row_df = pd.DataFrame([row], columns = self.headers)
                            df_81 = pd.concat([df_81, row_df], ignore_index=False)
                    elif id_91 != None and id_91 != "":
                        row = [id_91, name_91, "", prov, normal_id, element, month] + ["","",""] + ["","",""] + data_91
                        row_91_df = pd.DataFrame([row], columns = self.headers)
                        df_91 = pd.concat([df_91, row_91_df], ignore_index=False)
        if not df_81.empty: 
            df = pd.concat([df, df_81], ignore_index=False)      
        if not df_91.empty: 
            df = pd.concat([df, df_91], ignore_index=False)
        return df
                        