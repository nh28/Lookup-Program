import oracledb
import pandas as pd
import logging
import openpyxl

class ARKEON:
    def __init__(self):
        self.driver = 'oracle'
        self.ARKEON_host = 'ARC-CLUSTER.CMC.EC.GC.CA'
        self.ARKEON_port = '1521'
        self.ARKEON_service = 'archive.cmc.ec.gc.ca'
        self.connection = None
        try:
            self.workbook = openpyxl.load_workbook('StationList.xlsx')
            self.worksheet = self.workbook["Sheet1"]
        except Exception as e:
            print(e)
        
    def connect(self, usern, passw):
        """
        Connect to database

        Parameters:
        usern: Username that the user entered
        passw: Password that the user entered

        Returns:
        src_conn: The connection to the database
        """
        try:
            src_conn = oracledb.connect(
                user = usern,
                password = passw,
                host = self.ARKEON_host,
                port = self.ARKEON_port,
                service_name = self.ARKEON_service
            )
            self.connection = src_conn
            return True
        except oracledb.DatabaseError as err:
            error, = err.args
            logging.error('Unable to establish connection, due to: %s', error.message)
            return False
        
    def get_dataframe(self, column_names, from_section, where_section, where_used):
        """
        Turn the SQL query into a DataFrame

        Parameters:
        

        Returns:
        """
        select_clause = f'SELECT {", ".join(column_names)} FROM {from_section}'
        where_clause = ''
        if where_used:
            where_clause = f'WHERE {" AND ".join(where_section)}'

        query = f'{select_clause} {where_clause}'
        print(query)

        cursor = self.connection.cursor()

        if cursor is not None:
            logging.info('Executing query...')
            logging.info(query)
            try:
                cursor.execute(query)
                rows = cursor.fetchall()
            except Exception as err:
                msg = 'Unable to execute query, due to: %s' % str(err)
                logging.error(msg)
            logging.info('Query execution complete.')
        else:
            logging.error('Unable to execute query, due to: no cursor.')
            
        columns_data = {}
        for i, column_name in enumerate(column_names):
            columns_data[column_name] = [row[i] for row in rows]

        return pd.DataFrame(columns_data)
    
    def get_normals_element_id(self, element):
        normals = self.get_dataframe(['NORMAL_ID', 'E_NORMAL_ELEMENT_NAME'], 'NORMALS_1991.valid_normals_elements', [], False)

        for index, row in normals.iterrows():
            if element in row['E_NORMAL_ELEMENT_NAME']:
                return row['NORMAL_ID']
        return -1
    
    def get_extreme_el(self):
        all_el = self.get_all_elements()
        extremes = []
        for value in all_el:
            if "extreme" in value.lower():
                extremes.append(value)
        return extremes

    def get_all_stations(self):
        all_stations = []
        for row_idx in range(3, self.worksheet.max_row):
            name_7181 = self.worksheet.cell(row = row_idx, column = 2).value
            name_91 = self.worksheet.cell(row = row_idx, column = 7).value
            if name_7181 != None and name_7181 != "" and name_7181 not in all_stations:
                all_stations.append(name_7181)
            if name_91 != None and name_91 != "" and name_91 not in all_stations:
                all_stations.append(name_91)
        return sorted(all_stations)
    
    def get_all_elements(self):
        all_elements = []
        elements_1971 = self.get_dataframe(['e_normal_element_name'], 'NORMALS.valid_normals_elements', [], False) 
        elements_1981 = self.get_dataframe(['e_normal_element_name'], 'NORMALS_1981.valid_normals_elements', [], False)
        elements_1991 = self.get_dataframe(['e_normal_element_name'], 'NORMALS_1991.valid_normals_elements', [], False)

        for row in elements_1971.values:
            for value in row:
                if value != None and value not in all_elements:
                    all_elements.append(value)
        
        for row in elements_1981.values:
            for value in row:
                if value != None and value not in all_elements:
                    all_elements.append(value)
        
        for row in elements_1991.values:
            for value in row:
                if value != None and value not in all_elements:
                    all_elements.append(value)
        
        return sorted(all_elements)
